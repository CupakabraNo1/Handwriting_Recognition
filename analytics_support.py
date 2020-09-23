# -*- coding: utf-8 -*-
"""
Created on Mon Apr 26 09:14:08 2020

@author: CupakabraNo1
"""
# IMPORTS FOR CASE WHERE YOU DIDNT EXECUTE handewriting_recognition_main.py FIRST
# (DO NOT IMPORT THEM IF YOU ALREDY TRAINED YOUR MODEL)
# --------------------------------------------------------

from handwriting_recoginition_main import model, history
import constants as const
import dataset_loader as dl

# --------------------------------------------------------

import matplotlib.pyplot as plt

#----------------PLOT-ACCURACY----------------------------

plt.plot(range(1, const.EPOCHS+1), history.acc, 'green', label = 'train')
plt.plot(range(1, const.EPOCHS+1), history.val_acc, 'red', label = 'test')
plt.xlabel('Epochs')
plt.ylabel('Accuracy')
plt.legend()
plt.show()

#------------------PLOT-LOSS------------------------------

plt.plot(range(1, const.EPOCHS+1), history.loss, 'green', label = 'train')
plt.plot(range(1, const.EPOCHS+1), history.val_loss, 'red', label = 'test')
plt.xlabel('Epochs')
plt.ylabel('Loss')
plt.legend()
plt.show()

#--------------EVALUATE-ON-TEST-SET-----------------------

score = model.evaluate(dl.test_image_data, dl.test_label_data, batch_size = 1, verbose=0)
print('Test loss:', score[0])
print('Test accuracy:', score[1])

#---------------ACTIVATION-WAVES--------------------------

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side
from keras.models import Model
import numpy as np
import openpyxl as openpyxl

green_color = PatternFill(fgColor='F4FFE7', fill_type="solid")
green_color_strong = PatternFill(fgColor='2ECC71', fill_type="solid")
red_color = PatternFill(fgColor='FFE3E3', fill_type="solid")
red_color_strong = PatternFill(fgColor='E74C3C', fill_type="solid")

thin_border = Border(left=Side(style='thin', color='00000000'), 
                     right=Side(style='thin', color='00000000'), 
                     top=Side(style='thin', color='00000000'), 
                     bottom=Side(style='thin', color='00000000'))

accuracy_counter = 0
predictions = model.predict(dl.test_image_data[:const.SIZE_FOR_ANALISIS], batch_size = 1)

wb = ''

try:
    wb = openpyxl.load_workbook(model.name+const.EXT_CSV)
except :
    wb = Workbook()
    
for layer in model.layers:
    if(layer.name in const.LAYERS_FOR_ACT):

        layer_model = Model(inputs = model.input, outputs=model.get_layer(layer.name).output)
        layer_model_prediction = layer_model.predict(x = dl.test_image_data[:const.SIZE_FOR_ANALISIS], batch_size = 1)
        
        sheet = wb.create_sheet(layer.name)
        
        sheet.cell(row=1, column=1, value="REAL")
        sheet.cell(row=1, column=2, value="PREDICTED")
        
        for i in range(len(layer_model_prediction[0])):
            sheet.cell( row = 1, column = i+3, value=("N"+str(i)))
        
        for i in range(len(predictions)-1):
            row_num = 2+i;
            
            max_index_label = np.where(dl.test_label_data[i] == np.amax(dl.test_label_data[i]))
            max_index_prediction = np.where(predictions[i] == np.amax(predictions[i]))
            
            counter = 0;
            if(max_index_label[0][0] == max_index_prediction[0][0]):
                if(accuracy_counter == 0):
                    counter = counter + 1
                color = green_color
                
            else:
                color = red_color
            
            cell = sheet.cell(column=1, row=row_num, value=const.decode(max_index_label[0][0]))
            cell.fill = color
            cell.border = thin_border
            cell = sheet.cell(column=2, row=row_num, value=const.decode(max_index_prediction[0][0]))
            cell.fill = color
            cell.border = thin_border
            for j in range(len(layer_model_prediction[i])):
                cell = sheet.cell(column=3+j, row = row_num, value = layer_model_prediction[i][j])
                if(layer_model_prediction[i][j] > const.ACCTIVATION_BOUNDARY):
                    if(max_index_label[0][0] == max_index_prediction[0][0]):
                        color = green_color_strong
                    else: 
                        color = red_color_strong
                else:
                    if(max_index_label[0][0] == max_index_prediction[0][0]):
                        color = green_color
                    else: 
                        color = red_color
                cell.fill = color
                cell.border = thin_border
                
        if(accuracy_counter == 0 ): accuracy_counter = counter
        
wb.save(model.name+const.EXT_CSV)
print("Good predictions: " + str(accuracy_counter) + "/"+ str(len(predictions)))

